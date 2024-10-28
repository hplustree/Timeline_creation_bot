import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os

def csv_to_dataframe(csv_content):
    """
    Converts CSV content into a pandas DataFrame.
    
    Args:
        csv_content (str): CSV content as a string.
    
    Returns:
        DataFrame: A pandas DataFrame containing the CSV data.
    """
    csv_reader = csv.reader(csv_content.splitlines())
    columns = next(csv_reader)
    data = list(csv_reader)
    df = pd.DataFrame(data, columns=columns)

    # Convert 'Total Time (Days)' and 'Total Time (Hours)' to numeric
    df['Total Time (Days)'] = pd.to_numeric(df['Total Time (Days)'], errors='coerce', downcast='integer')
    df['Total Time (Hours)'] = pd.to_numeric(df['Total Time (Hours)'], errors='coerce', downcast='integer')

    return df

# Function to save DataFrame as Excel
def save_dataframe_to_excel(df, filename='project_timeline_temp.xlsx'):
    """
    Saves a pandas DataFrame to an Excel file.
    
    Args:
        df (DataFrame): A pandas DataFrame containing the project timeline.
        filename (str): The name of the Excel file to save.
    """
    df.to_excel(filename, index=False)

# Function to merge cells in Excel where values are identical
def merge_cells(ws, col_idx, df):
    """
    Merges cells in a given column where values are identical and aligns the text centrally.
    
    Args:
        ws (Worksheet): The active worksheet in the Excel file.
        col_idx (int): Column index (1-based) to apply the merge.
        df (DataFrame): The DataFrame to reference for row counting.
    """
    start_row = 2  # Excel index starts from 1, and header is in row 1
    for i in range(3, len(df) + 2 + 1):  # Adding 1 to ensure last row is included
        if i == len(df) + 2 or ws.cell(i, col_idx).value != ws.cell(i - 1, col_idx).value:
            if i - start_row > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=i - 1, end_column=col_idx)
            ws.cell(start_row, col_idx).alignment = Alignment(vertical='center', horizontal='center')
            start_row = i

# Function to adjust column width based on the longest text in the column
def auto_adjust_column_width(ws):
    """
    Automatically adjusts the width of each column based on the longest text in each column.
    
    Args:
        ws (Worksheet): The active worksheet in the Excel file.
    """
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (e.g., 'A', 'B')
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[column].width = adjusted_width
        
# Function to center align the entire column
def center_align_column(ws, col_idx, start_row, end_row):
    """
    Centers the alignment of all cells in a specified column.
    
    Args:
        ws (Worksheet): The active worksheet in the Excel file.
        col_idx (int): The index of the column to align (1-based).
        start_row (int): The starting row index.
        end_row (int): The ending row index.
    """
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=col_idx).alignment = Alignment(vertical='center', horizontal='center')

# Function to add a summary row at the end of the Excel file
def add_summary_row(ws, df):
    """
    Adds a summary row at the end of the worksheet with the total days and hours.
    
    Args:
        ws (Worksheet): The active worksheet in the Excel file.
        df (DataFrame): The DataFrame to calculate totals from.
    """
    total_days = df['Total Time (Days)'].sum()
    total_hours = df['Total Time (Hours)'].sum()
    
    # Determine the row number for the summary
    summary_row = len(df) + 2  # +1 for header and +1 for 1-based index

    # Insert the total values
    ws.cell(row=summary_row, column=1, value="Total").alignment = Alignment(vertical='center', horizontal='center')
    ws.cell(row=summary_row, column=2, value="Total").alignment = Alignment(vertical='center', horizontal='center')
    ws.cell(row=summary_row, column=3, value="Total").alignment = Alignment(vertical='center', horizontal='center')
    ws.cell(row=summary_row, column=4, value=total_days).alignment = Alignment(vertical='center', horizontal='center')
    ws.cell(row=summary_row, column=5, value=total_hours).alignment = Alignment(vertical='center', horizontal='center')

    # Merge the cells for the summary row
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)

# Main function to process the GPT response and generate an Excel file
def process_gpt_timeline_response(csv_content):
    """
    Processes GPT's timeline response, extracts CSV data, and generates an Excel file with merged cells,
    adjusted column widths, and developer side queries included.
    
    Args:
        csv_content (str): The GPT response containing CSV timeline data.
    """
    
    if csv_content:
        # Split the CSV content into sections based on empty lines
        sections = csv_content.split("\n\n")
        
        # Process the main timeline data
        timeline_data = sections[0]
        df = csv_to_dataframe(timeline_data)
        
        # Check if developer queries are provided
        developer_queries = sections[1] if len(sections) > 1 else None
        developer_queries_list = developer_queries.splitlines() if developer_queries else []

        # Remove the header "Developer Side Queries:" if it already exists in the list
        if developer_queries_list and developer_queries_list[0].strip().lower().startswith("developer side queries"):
            developer_queries_list.pop(0)

        # Save the DataFrame to a temporary Excel file
        save_dataframe_to_excel(df)

        # Load the Excel file to apply merging and column width adjustment
        wb = load_workbook('project_timeline_temp.xlsx')
        ws = wb.active

        # Define font and header color
        header_font = Font(name='Arial', bold=True, size=12, color="000000")  # Black font
        header_fill = PatternFill(start_color="89CFF0", end_color="89CFF0", fill_type="solid")  # Blue fill
        total_row_font = Font(name='Arial', bold=True, size=12, color="000000")  # Black font for totals
        total_row_fill = PatternFill(start_color="89CFF0", end_color="89CFF0", fill_type="solid")  # Blue fill for totals
        default_font = Font(name='Arial', size=11)  # Default font for all other cells

        # Define border style
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Merge cells for 'Phase', 'Task', and 'Subtask'
        merge_cells(ws, 1, df)  # Merge 'Phase' (Column 1 - A)
        merge_cells(ws, 2, df)  # Merge 'Task' (Column 2 - B)

        # Adjust column widths
        auto_adjust_column_width(ws)

        # Center align 'Total Time (Days)' and 'Total Time (Hours)' columns
        center_align_column(ws, 4, 2, len(df) + 1)  # Align 'Total Time (Days)' (Column 4 - D)
        center_align_column(ws, 5, 2, len(df) + 1)  # Align 'Total Time (Hours)' (Column 5 - E)

        # Add summary row with totals
        add_summary_row(ws, df)

        # Apply Arial font to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.font = default_font  # Set default font for all cells
                cell.border = thin_border  # Add border to header cells

        # Set header styles
        for cell in ws[1]:  # Assuming headers are in the first row
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border  # Add border to header cells

        # Style the last row (summary totals)
        last_row = len(df) + 2  # Assuming summary is the last row
        for cell in ws[last_row]:  # Apply styles to the last row
            cell.font = total_row_font
            cell.fill = total_row_fill
            cell.border = thin_border  # Add border to header cells

        # Insert the Developer Side Queries section below the Total Time table
        if developer_queries_list:
            # Leave a blank row after the summary
            dev_query_start_row = last_row + 2
            ws.cell(row=dev_query_start_row, column=1, value="Developer Side Queries:")
            ws.cell(row=dev_query_start_row, column=1).font = header_font
            ws.cell(row=dev_query_start_row, column=1).alignment = Alignment(vertical='center', horizontal='left')

            # Add each query to a new row
            for idx, query in enumerate(developer_queries_list, start=dev_query_start_row + 1):
                ws.cell(row=idx, column=1, value=query)
                ws.cell(row=idx, column=1).alignment = Alignment(vertical='top', horizontal='left')
                ws.cell(row=idx, column=1).font = default_font

        # Save the final Excel file with merged cells, adjusted widths, and developer queries
        wb.save('project_timeline.xlsx')

        # Remove the temp excel
        if os.path.exists('project_timeline_temp.xlsx'):
            os.remove('project_timeline_temp.xlsx')

        print("Timeline saved as 'project_timeline.xlsx' with merged cells, adjusted column widths, and developer queries.")
        return 'project_timeline.xlsx'
    else:
        print("No CSV content found in the GPT response.")
